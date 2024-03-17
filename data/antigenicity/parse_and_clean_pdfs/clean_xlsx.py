import openpyxl
import csv, os, sys
import csv

input_dir = sys.argv[1] #  "dir/of/your/reports"
save_root_dir = sys.argv[2] # "dir/of/clean/tables"
subtype = sys.argv[3] # h3n2 or h1n1
save_root_dir = os.path.join(save_root_dir, subtype.replace(" ", "_").lower())

if not os.path.exists(save_root_dir):
    os.makedirs(save_root_dir)

def read_xlsx(path):
    wb = openpyxl.load_workbook(filename=path, data_only=True)
    sheet_obj = wb.active
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column

    table = []
    for i in range(1, max_row + 1):
        row_values = [sheet_obj.cell(row = i, column = j).value for j in range(1, max_col + 1)]
        row_values = [x.lower() if isinstance(x, str) else x for x in row_values]
        table.append(row_values)
    return table, max_row, max_col

def read_csv(path):
    table, max_row, max_col = [], 0, 0
    with open(path) as csvfile:
        spamreader = csv.reader(csvfile)
        for row in spamreader:
            max_col = max(max_col, len(row))
            max_row += 1
            table.append([x.lower() for x in row])
    return table, max_row, max_col

def is_virus_name(cell):
    if isinstance(cell, str):
        if len(cell.split("/")) >= 3:
            return True
        else:
            return False
    else:
        return False

def is_valid_hi(cell):
    # float/int, or <40, <, nd
    if isinstance(cell, float) or isinstance(cell, int):
        return True

    if cell == "<" or cell == "nd":
        return True

    if cell is not None and len(cell) > 0 and cell[0] != "<":
        try:
            cell = float(cell)
            return True
        except Exception as e:
            return False
    
    if cell is not None and len(cell) > 0 and cell[0] == "<":
        try:
            cell = float(cell[1:])
            return True
        except Exception as e:
            return False

    return False
    
    # if isinstance(cell, float) or isinstance(cell, int) or cell == "<" or cell == "nd":
    #     return True
    # else:
    #     return False

error_tables = []

for file in os.listdir(input_dir):
    if file.endswith('.xlsx') or file.endswith('.csv'):
        path = os.path.join(input_dir, file)
        print("Extracting tables from", path)

        if path.endswith('.xlsx'):
            table, max_row, max_col = read_xlsx(path)
        elif path.endswith('.csv'):
            table, max_row, max_col = read_csv(path)

        antigenic_tables_headline = []
        for i in range(max_row):
            for j in range(max_col):
                if isinstance(table[i][j], str) and "table" in table[i][j] and "antigenic analyses" in table[i][j]:
                    antigenic_tables_headline.append(i)

        antigenic_tables_headline.sort()
        # print(antigenic_tables_headline)

        antigenic_tables_content = {}
        for headline_row in antigenic_tables_headline:
            # finished_table = False
            # Viruses 
            viruses_col = None
            info_row = -1
            start_of_content_row = -1
            first_split_row = -1
            row_with_content = {}
            for row in range(headline_row, max_row):
                if "viruses" in table[row]:
                    viruses_col = table[row].index("viruses")
                    # other_possible_virus_cols = [] # [viruses_col + 1 if table[row][viruses_col + 1] is None]
                    # if table[row][viruses_col + 1] is None:
                    #     other_possible_virus_cols.append(viruses_col + 1)
                    # if viruses_col - 1 >=0 and table[row][viruses_col - 1] is None:
                    #     other_possible_virus_cols.append(viruses_col - 1)
                    info_row = row
                
                if sum([int(is_virus_name(x)) for x in table[row]]) > 0:
                    row_with_content[row] = True
                else:
                    row_with_content[row] = False

                
                if start_of_content_row == -1 and viruses_col is not None and is_virus_name(table[row][viruses_col]):
                    start_of_content_row = row
                
                # if first_split_row == -1 and start_of_content_row > -1 and not is_virus_name(table[row][viruses_col]) \
                #     and row-1 >=0 and is_virus_name(table[row-1][viruses_col]) and row+1 < max_row \
                #         and is_virus_name(table[row+1][viruses_col]):
                #     first_split_row = row

                # # The current row has the content, while the 
                # if first_split_row == -1 and row_with_content[row] and not row_with_content[row-1] and row_with_content[row-2]:
                #     first_split_row = row - 1 
                
                if viruses_col is not None:
                    # two concecutive rows not including any contents
                    if start_of_content_row > -1 and not row_with_content[row] and not row_with_content[row-1]:
                        break
                    # this cell is not a virus name, but previous cell is a virus name, the next cell is not a virus name as well.
                    # if is_virus_name(table[row-1][viruses_col]) and not is_virus_name(table[row][viruses_col]) and not is_virus_name(table[row+1][viruses_col]):
                        # break
                    
                    # if not is_virus_name(table[row][viruses_col]) and (table[row][viruses_col] is None or table[row+1][viruses_col] is None) and is_virus_name(table[row-1][viruses_col]):
                        # break

            finished_table_row = row - 1
            # print(headline_row, finished_table_row, info_row)
            table_content = []

            cols_with_hi_value = []
            # empty_cols = []
            valid_cols = []
            for col in range(len(table[headline_row])):
                cols_value = [table[i][col] for i in range(info_row, finished_table_row)] # 要么是flaot value要么是<要么是ND？？
                if "genetic group" in cols_value:
                    continue
                elif "genetic" in cols_value: # might be mistakely regarded as a HI column
                    continue
                    # exit()
                # print(cols_value, sum([int(is_valid_hi(x)) for x in cols_value]))
                # print(sum([int(is_valid_hi(x)) for x in cols_value]))
                if sum([int(is_valid_hi(x)) for x in cols_value]) > len(cols_value) * 0.4:
                    valid_cols.append(col)
            
            for c in table[headline_row]:
                if isinstance(c, str) and ( "table" in c or "tab" in c):
                    table_name = c
            
            if subtype not in table_name:
                continue


            info_row_value = ["viruses"] + [table[info_row][col] for col in valid_cols]
            # Try to fix mistake caused by pdf parsing or messy data
            info_row_value = [x.replace("2a/", " a/").strip() if x is not None else x for x in info_row_value]
            if len(info_row_value) > 1 and info_row_value[1] is not None and "passage" in info_row_value[1]: # # Passage xxx
                info_row_value = ["viruses"] + [" ".join(info_row_value[1].split("passage")[1:]).strip()] + info_row_value[2:]

            if sum([int(x is None) for x in info_row_value]) > 0:
                info_row_value_resplit = ["viruses"]
                for x in info_row_value[1:]:
                    if x is not None:
                        x_split = []
                        for part in x.split():
                            if part.startswith("a/") or part.startswith("nymc") or part.startswith("sw") or part.startswith("x-147") or part.startswith("ivr-"): 
                                x_split.append(part)
                            else:
                                # print(x_split, part, info_row_value)
                                x_split[-1] = x_split[-1] + " " + part
                        info_row_value_resplit.extend(x_split)
                if len(info_row_value_resplit) == len(info_row_value):
                    # print("resplit", info_row_value, info_row_value_resplit)
                    info_row_value = info_row_value_resplit


            info_row_value_extra = ["viruses"] + [table[info_row+1][col] for col in valid_cols]
            # print(info_row_value)
            # print(info_row_value_extra)
            error = False
            for i in range(1, len(info_row_value)):
                if info_row_value[i] is None:
                    error_tables.append((file, table_name, info_row_value))
                    error = True
                    continue

                if info_row_value_extra[i] is not None:
                    if len(str(info_row_value_extra[i])) > 0:
                        info_row_value[i] = info_row_value[i] + "/" + str(info_row_value_extra[i])
            if error:
                continue
            
            table_content.append(info_row_value)
            for i in range(headline_row, finished_table_row):
                virus_name = None
                for j in range(len(table[i])):
                    if is_virus_name(table[i][j]):
                        virus_name = table[i][j]
                        break
                if virus_name is None:
                    continue
                    
                if sum([is_valid_hi(table[i][j])  for j in valid_cols]) > len(valid_cols) * 0.5:
                    table_content.append([virus_name] + [table[i][j] for j in valid_cols])
                    # print([virus_name] + [table[i][j] for j in valid_cols])
            
            ref_virus = [x[0] for x in table_content[1:]]
            antigenic_tables_content[table_name] = table_content
            # print(info_row_value)

        # print(len(antigenic_tables_content))
        # print(antigenic_tables_content[0])
        save_dir = os.path.join(save_root_dir, os.path.split(path)[1].replace(" ", "_").split(".xlsx")[0])

        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        for name in antigenic_tables_content:
            save_name = name.replace(" ", "_").replace(".", "").replace("/", "_")
            # print("Saving %s" % os.path.join(save_dir, save_name + ".csv"))
            with open(os.path.join(save_dir, save_name + ".csv"), "w") as csvfile:
                spamwriter = csv.writer(csvfile)
                for row in antigenic_tables_content[name]:
                    spamwriter.writerow(row)
                # spamwriter.writerow(['Spam', 'Lovely Spam', 'Wonderful Spam'])

if len(error_tables) > 0:
    print("Fail to extract info from these tables: (please check manually)")
    for error_table in error_tables:
        print(error_table)